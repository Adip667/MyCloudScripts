import boto3
from time import strftime
import configparser
from botocore.exceptions import ClientError, WaiterError
from openpyxl import Workbook, load_workbook
import argparse


def get_config_regions():
    """
    read the region configuration from config.txt
    :return: return list of region based on user config.txt
    """
    existing_regions = ('eu-north-1', 'ap-south-1', 'eu-west-3', 'eu-west-2', 'eu-west-1', 'ap-northeast-2',
                        'ap-northeast-1', 'sa-east-1', 'ca-central-1', 'ap-southeast-1', 'ap-southeast-2',
                        'eu-central-1', 'us-east-1', 'us-east-2', 'us-west-1', 'us-west-2')

    _log('INFO: Checking region config')
    config = configparser.ConfigParser()
    config.read('config.txt')
    region_list = []

    if config['ec2_region'].getboolean('All'):
        _log('INFO: Regions from config file are - All regions')
        return existing_regions
    else:
        region_list = config['ec2_region']['regions'].split(",")
        bad_region = [region for region in region_list if region.strip() not in existing_regions]
        if bad_region:
            _log(f"ERROR: Not found - {bad_region}, Please check your configuration")
        region_list = [region for region in region_list if region.strip() in existing_regions]
        _log(f"INFO: Valid regions from config file are - {region_list}")
        return region_list


def get_config_account():
    """
    get the account number from the config file, used for snapshot and images
    :return: account number in list.
    """
    _log('INFO: Checking account config')
    config = configparser.ConfigParser()
    config.read('config.txt')
    account_details = [config['aws_details']['aws_account']]
    _log(f"INFO: Found account: {account_details[0]}")
    return account_details


def clean_ec2(dry_run=True):
    _log("INFO: Starting EC2 cleaning")

    # going over each region configured and checking for EC2
    for region in regions:
        _log(f"INFO: Checking EC2 instances in region - {region}")

        stop_list = []  # will store list of EC2 to be shutdown
        terminate_list = []  # will store list of EC2 to be terminated

        ec2 = boto3.client('ec2', region_name=region.strip())
        response = ec2.describe_instances()
        ec2_instances = [i for instance in response['Reservations'] for i in
                         instance['Instances']]  # extract the list of instances from the response
        if not ec2_instances:
            _log(f'WARNING: region {region}: No EC2 instances found')
        else:
            _log(f'INFO: region {region}: Found EC2 instances')
            for instance in ec2_instances:

                # get tags and check what operation need to be done
                if not instance.get('Tags'):
                    Tags = 'N/A'
                    operation = 'Terminate'
                else:
                    Tags = {tag.get('Key'): tag.get('Value') for tag in instance.get('Tags')}
                    if Tags.get('keep') == 'on':
                        operation = 'DoNothing'
                    elif Tags.get('keep') == 'off':
                        operation = 'Shutdown'
                    else:
                        operation = 'Terminate'

                _log(f"INFO: instance: {instance}")
                print_results_xlsx(data=instance, sheetname='EC2', Tags=str(Tags), OperationDone=operation)

                if operation == 'Shutdown':
                    stop_list.append(instance['InstanceId'])
                elif operation == 'Terminate':
                    terminate_list.append(instance['InstanceId'])

            if stop_list:  # stop the instances
                _log(f'INFO: Stopping in region{region}: {stop_list}')
                try:
                    response = ec2.stop_instances(InstanceIds=stop_list, DryRun=dry_run)
                    _log(f"INFO: Stopping instance response {response}")
                except ClientError as e:
                    _log(f"ERROR: {e}")

                    print_results_xlsx(data=str(stop_list), sheetname='EC2', OperationDone='ERROR-Shutdown',
                                       error=str(e))

            if terminate_list:  # terminate the instances
                _log(f'INFO: Terminating in region{region}: {terminate_list}')
                try:
                    response = ec2.terminate_instances(InstanceIds=terminate_list, DryRun=dry_run)
                    _log(f"INFO: terminate instance response {response}")
                except ClientError as e:  # probably some permission error
                    _log(f"ERROR: {e}")
                    print_results_xlsx(data=str(terminate_list), sheetname='EC2', OperationDone='ERROR-Terminate',
                                       error=str(e))
                else:  # if termination raised no error, check if it finished (as volume are depended on this)
                    try:
                        waiter = ec2.get_waiter('instance_terminated')
                        waiter.wait(InstanceIds=terminate_list, WaiterConfig={'Delay': 15, 'MaxAttempts': 12},
                                    DryRun=dry_run)
                    except WaiterError as e:
                        _log(f"ERROR: {e}")
                        print_results_xlsx(data=str(terminate_list), sheetname='EC2',
                                           OperationDone='ERROR-waitTerminate',
                                           error=str(e))

        _log(f"INFO: region end: {region}")
    _log("INFO: existing clean_ec2()")


def clean_snapshot(dry_run=True):
    """
    check for snapshot in all region and delete them all
    :param dry_run: for BOTO3 call
    """
    _log("INFO: entering clean_snapshot()")
    account = get_config_account()
    for region in regions:
        _log(f'INFO: Cleaning all snapshots for {region}')
        ec2 = boto3.client('ec2', region_name=region.strip())
        response = ec2.describe_snapshots(OwnerIds=account)
        _log(f'describe_snapshots response {response}')
        for snap in response['Snapshots']:
            try:
                _log(f"INFO: Found {snap['SnapshotId']} for volume: {snap['VolumeId']}, size {snap['VolumeSize']} GB")
                ec2.delete_snapshot(SnapshotId=snap['SnapshotId'], DryRun=dry_run)
            except ClientError as e:
                _log(f'ERROR: {e}')
                print_results_xlsx(data=snap, sheetname='Snapshots', region=region, error=e)
            else:
                print_results_xlsx(data=snap, sheetname='Snapshots', region=region)
    _log("INFO: existing clean_snapshot()")


def clean_volumes(dry_run=True):
    """
    check for volumes in all regions and delete all state=available volumes
    :param dry_run: for BOTO 3 call
    """
    _log("INFO: entering clean_volumes()")
    for region in regions:
        _log(f'INFO: Cleaning available volumes for {region}')
        ec2 = boto3.client('ec2', region_name=region.strip())
        response = ec2.describe_volumes()
        for volume in response['Volumes']:
            try:

                Tags = volume.get('Tags')
                if Tags:
                    Tags = {tag.get('Key'): tag.get('Value') for tag in Tags}
                _log(
                    f"INFO: Found volume in {volume['AvailabilityZone']}: {volume['VolumeId']}({volume['State']},"
                    f" {volume['Iops']} IOPS, {volume['VolumeType']}) with Tag: {Tags}"
                )
                state = 'Nothing'

                if volume['State'] == 'available':
                    state = 'Terminate'
                    _log('INFO: Deleting Volume')
                    ec2.delete_volume(VolumeId=volume['VolumeId'], DryRun=dry_run)
            except ClientError as e:
                _log(f'ERROR: {e}')
                print_results_xlsx(data=volume, sheetname='Volumes', Tags=Tags, OperationDone=state, error=e)
            else:
                print_results_xlsx(data=volume, sheetname='Volumes', Tags=Tags, OperationDone=state)

    _log("INFO: existing clean_volumes()")


def clean_images(dry_run=True):
    """
    check for AMI's in all regions and Deregister if there is no tag keep
    :param dry_run: for BOTO 3 call
    """

    _log("INFO: entering clean_images()")
    account = get_config_account()
    for region in regions:
        _log(f'INFO: Cleaning available images for {region}')
        ec2 = boto3.client('ec2', region_name=region.strip())
        images = ec2.describe_images(Owners=account)
        if not images['Images']:
            _log(f'WARNING: no images found for {region}')
        else:
            for img in images['Images']:
                try:
                    OperationDone = ''
                    Tags = img.get('Tags')
                    if Tags:
                        Tags = {tag.get('Key'): tag.get('Value') for tag in Tags}

                        if 'keep' in Tags:
                            OperationDone = "Keep"
                        else:

                            OperationDone = "Deregister"
                            ec2.deregister_image(ImageId=img['ImageId'], DryRun=dry_run)

                    else:

                        OperationDone = "Deregister"
                        ec2.deregister_image(ImageId=img['ImageId'], DryRun=dry_run)

                except ClientError as e:
                    print_results_xlsx(data=img, sheetname='Images', region=region, OperationDone=OperationDone,
                                       Tags=Tags, error=e)
                else:
                    print_results_xlsx(data=img, sheetname='Images', region=region, OperationDone=OperationDone,
                                       Tags=Tags)
    _log("INFO: existing clean_images()")


def clean_sg(dry_run=True):
    """
    Check each region for security groups with boto3, delete SG that are unused & untagged
    :param dry_run: used for boto call, to avoid actually deleting anything
    :return: None
    """
    _log(f"INFO: Cleaning SG")
    headers = ["Region", "OwnerId", "SG Name", "SG Id", "VpcId", "FromPort",
               "ToPort", "IpProtocol", "Source", "Instances", "Tags", "OperationDone"]

    for region in regions:  # iterate over the region list and get the SG's
        ec2 = boto3.client('ec2', region_name=region.strip())
        response = ec2.describe_security_groups()

        _log(f"INFO: Checking SG in region - {region}")

        security_group_record = {'Region': region}  # dict for the SG, will be send later to the report

        for sg in response['SecurityGroups']:  # iterate over all the SG in the current region and add data to dict
            _log(f"INFO: Found security group")
            _log(f"INFO: {sg}")

            security_group_record['GroupName'] = sg['GroupName']
            security_group_record['VpcId'] = sg.get('VpcId')
            security_group_record['OwnerId'] = sg.get('OwnerId')

            security_group_record['Instances'] = ''

            # get instances so we have SG -> relation
            instances_for_sg = ec2.describe_instances(
                Filters=[{'Name': 'instance.group-id', 'Values': [sg.get('GroupId'), ]}, ])
            instances_for_sg = [i for instance in instances_for_sg['Reservations'] for i in
                                instance['Instances']]
            instances_for_sg = [instance['InstanceId'] for instance in instances_for_sg]

            # set OperationDone to N/A, will be updated later if we delete
            security_group_record['OperationDone'] = 'N/A'
            security_group_record['GroupId'] = sg.get('GroupId')

            delete_error = "N/A"
            if not instances_for_sg:  # if no instances found, check for tag and update 'OperationDone'
                security_group_record['Instances'] = 'N/A'

                sg_tag_no_delete = False

                if sg.get('Tags'):  # check if there are any tags at all
                    for tag in sg.get('Tags'):  # check for the relevant tag
                        if tag.get('Key') == 'keep':
                            _log('INFO: Found no delete tag(keep)')
                            sg_tag_no_delete = True  # don't delete
                if not sg_tag_no_delete:
                    security_group_record['OperationDone'] = 'Deleting'
                    _log(f'INFO: removing sg - {sg.get("GroupId")}')
                    try:
                        ec2.delete_security_group(GroupId=sg.get('GroupId'), DryRun=dry_run)
                    except ClientError as e:
                        print(f"\tERROR: {str(e.response['Error'])}")
                        delete_error = e

            else:
                security_group_record['Instances'] = ', '.join(instances_for_sg)  # convert instance list to string

            print_results_xlsx(data=security_group_record, sheetname='SG',
                               OperationDone=security_group_record['OperationDone'], error=delete_error)
        _log("INFO: Region END")


def create_xlsx():
    _log('INFO: Creating excel')
    wb = Workbook()

    ws_ec2 = wb.active
    ws_ec2.title = 'EC2'
    ws_ec2.append(
        ("OperationDone", "InstanceId", "InstanceType", "AvailabilityZone", "PrivateIpAddress", "PublicDnsName",
         "State", "SubnetId", "VpcId", "RootDeviceType", "Volumes", "SecurityGroups Name", "SecurityGroups",
         "Tags"))

    ws_volumes = wb.create_sheet()
    ws_volumes.title = 'Volumes'
    ws_volumes.append(
        ("OperationDone", "VolumeId", "AvailabilityZone", "State", "Iops", "VolumeType", "Tags", "Errors"))

    ws_snapshots = wb.create_sheet()
    ws_snapshots.title = 'Snapshots'
    ws_snapshots.append(("SnapshotID(deleted)", "VolumeId", "Region", "Errors"))

    ws_images = wb.create_sheet()
    ws_images.title = 'Images'
    ws_images.append(
        ("OperationDone", "ImageId", "Name", "Region", "OwnerId", "ImageType", "CreationDate", "Tags", "Errors"))

    ws_images = wb.create_sheet()
    ws_images.title = 'SG'
    ws_images.append(
        ("OperationDone", "SG Id", "SG Name", "OwnerId", "Region", "VpcId", "Instances", "Errors"))

    wb.save(xlsx_name)


def print_results_xlsx(**kwargs):
    wb = load_workbook(xlsx_name)
    ws = wb[kwargs['sheetname']]

    error = kwargs.get('error')
    if kwargs['sheetname'] == 'Volumes':
        row = (
            kwargs['OperationDone'], kwargs['data']['VolumeId'], kwargs['data']['AvailabilityZone'],
            kwargs['data']['State'], kwargs['data']['Iops'], kwargs['data']['VolumeType'],
            str(kwargs['Tags']), str(error)
        )
        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'Snapshots':
        row = (kwargs['data']['SnapshotId'], kwargs['data']['VolumeId'], kwargs['region'], str(error))
        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'Images':
        row = (kwargs['OperationDone'], kwargs['data']["ImageId"], kwargs['data']["Name"], kwargs['region'],
               kwargs['data']["OwnerId"], kwargs['data']["ImageType"], kwargs['data']["CreationDate"],
               str(kwargs["Tags"]), str(error))
        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'EC2' and error == None:

        if not kwargs['data']['PublicDnsName']: kwargs['data']['PublicDnsName'] = 'N/A'

        volume_list = ''
        for volume in kwargs['data']['BlockDeviceMappings']:
            volume_list += f"{(volume['Ebs']['VolumeId'])}({volume['Ebs']['Status']}),  "

        sg_list_name = ''
        sg_list_id = ''
        for sg in kwargs['data']['SecurityGroups']:
            sg_list_name += f"{sg['GroupName']},  "
            sg_list_id += f"{sg['GroupId']},  "

        row = (kwargs['OperationDone'], kwargs['data']['InstanceId'], kwargs['data']['InstanceType'],
               kwargs['data']['Placement']['AvailabilityZone'],
               kwargs['data'].get('PrivateIpAddress'), kwargs['data']['PublicDnsName'], kwargs['data']['State']['Name'],
               kwargs['data'].get('SubnetId'),
               kwargs['data'].get('VpcId'), kwargs['data']['RootDeviceType'], volume_list, sg_list_name, sg_list_id,
               kwargs['Tags'])
        ws.append(row)
        wb.save(xlsx_name)
    elif kwargs['sheetname'] == 'EC2':
        ws.append((kwargs['OperationDone'], kwargs['data'], error))
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'SG':

        row = (
        kwargs['OperationDone'], kwargs['data']["GroupId"], kwargs['data']["GroupName"], kwargs['data']["OwnerId"],
        kwargs['data']['Region'], kwargs['data']["VpcId"], kwargs['data']["Instances"],
        str(error))

        ws.append(row)
        wb.save(xlsx_name)


def _log(line):
    console = True

    if Logfile:
        with open(log_name, "a") as file:
            file.write(str(line) + '\n')
    if console:
        print(line)


if __name__ == '__main__':

    dryrun = False
    Logfile = False

    parser = argparse.ArgumentParser(description='Run cleanup as config in the config.txt file')
    parser.add_argument('--operation', '-o', type=str,
                        help='an operation name, Can be "storage" (EC2/Images/Snapshots/Volume) or "sg" (security groups)')
    parser.add_argument('--dryrun', metavar='Bool', type=str,
                        help='Run in dry run mode, wont delete anything if set to True')
    parser.add_argument('--log', metavar='Bool', type=str,
                        help='Will create logs file for the CLI Operations')

    args = parser.parse_args()

    log_name = strftime('clean_log_' + "%Y-%b-%d_%H-%M-%S.log")
    xlsx_name = strftime('ServiceCleaner_' + "%Y-%b-%d_%H-%M-%S.xlsx")
    regions = get_config_regions()

    if (args.log == 'True'):
        Logfile = True

    if (args.dryrun == 'True'):
        dryrun = True

    if (args.operation == 'storage'):
        _log(f"INFO: Cleaning storage")
        create_xlsx()
        clean_ec2(dryrun)
        clean_volumes(dryrun)
        clean_images(dryrun)
        clean_snapshot(dryrun)


    elif (args.operation == 'sg'):
        _log(f"INFO: Cleaning Security Groups")
        create_xlsx()
        clean_sg(dryrun)


    elif (args.operation == 'all'):
        _log(f"INFO: Cleaning Storage and SG")
        create_xlsx()
        clean_ec2(dryrun)
        clean_volumes(dryrun)
        clean_images(dryrun)
        clean_snapshot(dryrun)
        clean_sg(dryrun)

    else:
        _log(f"INFO: provided argument is incorrect:\n  operation={args.operation}")
